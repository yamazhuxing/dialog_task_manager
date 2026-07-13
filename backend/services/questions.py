import io
import json
import zipfile
from datetime import datetime, timezone
from pathlib import Path

from sqlalchemy import func
from sqlalchemy.orm import Session

from backend.config import Settings
from backend.constants import MAX_TASK_TURNS, MIN_TASK_TURNS, SCENE_LABELS
from backend.models import QuestionImport, Sample, Task, User
from backend.services.quality_report import remove_convert_metadata
from backend.services.sample_paths import SourceSamplePaths, iter_delivery_sources
from backend.services.submission_validation import sample_storage_dir_name

DELIVERY_ZIP_NAME = "delivery_latest.zip"
DELIVERY_ZIP_META_NAME = "delivery_latest.meta.json"
RAW_DELIVERY_ZIP_NAME = "delivery_raw_latest.zip"
RAW_DELIVERY_ZIP_META_NAME = "delivery_raw_latest.meta.json"
V2_DELIVERY_ZIP_NAME = "delivery_v2_latest.zip"
V2_DELIVERY_ZIP_META_NAME = "delivery_v2_latest.meta.json"
RAW_MANIFEST_NAME = "raw_manifest.json"
QC_RECORD_XLSX_NAME = "质检提交记录.xlsx"


def import_questions_from_data(
    db: Session,
    questions: list[dict],
    *,
    filename: str,
    imported_by: User,
    import_batch: str | None = None,
) -> tuple[int, int]:
    imported = 0
    skipped = 0
    batch = import_batch or datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")

    for item in questions:
        task_id = int(item["id"])
        exists = db.get(Task, task_id)
        if exists:
            skipped += 1
            continue
        task = Task(
            id=task_id,
            scene=item["scene"],
            scene_label=item.get("scene_label", item["scene"]),
            topic=item.get("topic", ""),
            constraint_text=item.get("constraint"),
            turns_json=json.dumps(item.get("turns", []), ensure_ascii=False),
            design_notes_json=json.dumps(item.get("design_notes"), ensure_ascii=False)
            if item.get("design_notes")
            else None,
            import_batch=batch,
        )
        db.add(task)
        imported += 1

    record = QuestionImport(
        filename=filename,
        imported_count=imported,
        skipped_count=skipped,
        imported_by_id=imported_by.id,
    )
    db.add(record)
    db.commit()
    return imported, skipped


def load_questions_file(path: Path) -> list[dict]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError("题目文件必须是 JSON 数组")
    return data


def create_single_task(db: Session, *, scene: str, topic: str, constraint_text: str | None, turns: list[dict], created_by: User) -> Task:
    scene_label = SCENE_LABELS.get(scene)
    if not scene_label:
        raise ValueError("无效的场景类型")

    if not (MIN_TASK_TURNS <= len(turns) <= MAX_TASK_TURNS):
        raise ValueError(f"提问轮数需在 {MIN_TASK_TURNS}~{MAX_TASK_TURNS} 之间")

    normalized_turns = []
    for idx, turn in enumerate(turns, start=1):
        content = (turn.get("content") or "").strip()
        if not content:
            raise ValueError(f"第 {idx} 轮提问不能为空")
        normalized_turns.append({"round": idx, "role": "user", "content": content})

    max_id = db.query(func.max(Task.id)).scalar() or 0
    task = Task(
        id=max_id + 1,
        scene=scene,
        scene_label=scene_label,
        topic=topic.strip(),
        constraint_text=constraint_text.strip() if constraint_text else None,
        turns_json=json.dumps(normalized_turns, ensure_ascii=False),
        design_notes_json=json.dumps(
            {
                "rounds": len(normalized_turns),
                "generator": "manual",
                "created_by": created_by.username,
            },
            ensure_ascii=False,
        ),
        import_batch="manual",
    )
    db.add(task)
    db.commit()
    db.refresh(task)
    return task


def _delivery_fingerprint(settings: Settings) -> tuple[int, float, list[tuple[str, Path]]]:
    """统计交付目录文件数、最新修改时间，并收集待打包目录。"""
    delivery_sources = iter_delivery_sources(settings.samples_dir)
    if not delivery_sources:
        raise FileNotFoundError("暂无已通过样本，请先完成至少一条通过样本")

    zip_entries: list[tuple[str, Path]] = []
    file_count = 0
    max_mtime = 0.0
    for paths in delivery_sources:
        if not paths.convert_dir.exists() or not paths.qc_dir.exists():
            continue
        remove_convert_metadata(paths.convert_dir)
        zip_entries.append((paths.convert_dir.name, paths.convert_dir))
        zip_entries.append((paths.qc_dir.name, paths.qc_dir))

    if not zip_entries:
        raise FileNotFoundError("暂无可交付的样本目录，请先完成至少一条通过样本")

    for _root_name, folder in zip_entries:
        for file_path in folder.rglob("*"):
            if file_path.is_file():
                file_count += 1
                max_mtime = max(max_mtime, file_path.stat().st_mtime)
    return file_count, max_mtime, zip_entries


def _read_delivery_cache_meta(meta_path: Path) -> dict | None:
    if not meta_path.exists():
        return None
    try:
        data = json.loads(meta_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    return data if isinstance(data, dict) else None


def _write_delivery_zip(settings: Settings, zip_entries: list[tuple[str, Path]], zip_path: Path) -> None:
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED, compresslevel=1) as zf:
        for root_name, folder in zip_entries:
            for file_path in folder.rglob("*"):
                if file_path.is_file():
                    arcname = f"{root_name}/{file_path.relative_to(folder).as_posix()}"
                    zf.write(file_path, arcname)


def create_delivery_zip(settings: Settings, *, force: bool = False) -> Path:
    file_count, max_mtime, zip_entries = _delivery_fingerprint(settings)
    cache_zip = settings.data_dir / DELIVERY_ZIP_NAME
    cache_meta = settings.data_dir / DELIVERY_ZIP_META_NAME

    cached = _read_delivery_cache_meta(cache_meta)
    if (
        not force
        and cache_zip.exists()
        and cached
        and cached.get("file_count") == file_count
        and cached.get("max_mtime") == max_mtime
    ):
        return cache_zip

    _write_delivery_zip(settings, zip_entries, cache_zip)
    cache_meta.write_text(
        json.dumps(
            {
                "file_count": file_count,
                "max_mtime": max_mtime,
                "built_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return cache_zip


def _resolve_sample_raw_path(settings: Settings, sample: Sample) -> Path | None:
    """定位已通过样本的原始上传文件。"""
    candidates: list[Path] = []
    stored = Path(sample.raw_file_path)
    candidates.append(stored)
    if not stored.is_absolute():
        candidates.append(settings.project_root / stored)
    candidates.append(settings.samples_dir / sample.source_type / stored.name)
    for path in candidates:
        if path.is_file():
            return path
    return None


def _raw_delivery_fingerprint(db: Session, settings: Settings) -> tuple[int, float, list[dict]]:
    """统计原始文件数、最新修改时间，并收集待打包条目。"""
    rows = (
        db.query(Sample, User.username)
        .join(User, Sample.user_id == User.id)
        .order_by(Sample.id.asc())
        .all()
    )
    if not rows:
        raise FileNotFoundError("暂无已通过样本，请先完成至少一条通过样本")

    entries: list[dict] = []
    file_count = 0
    max_mtime = 0.0
    for sample, username in rows:
        raw_path = _resolve_sample_raw_path(settings, sample)
        if raw_path is None:
            continue
        file_count += 1
        max_mtime = max(max_mtime, raw_path.stat().st_mtime)
        entries.append(
            {
                "arcname": f"{sample.source_type}/{raw_path.name}",
                "path": raw_path,
                "manifest": {
                    "task_id": sample.task_id,
                    "session_id": sample.session_id,
                    "user_id": sample.user_id,
                    "username": username,
                    "source_type": sample.source_type,
                    "model_version": sample.model_version,
                    "scene": sample.scene,
                    "difficulty": sample.difficulty,
                    "arcname": f"{sample.source_type}/{raw_path.name}",
                },
            }
        )

    if not entries:
        raise FileNotFoundError("暂无可下载的原始上传文件，请检查 samples 表中的 raw_file_path")

    return file_count, max_mtime, entries


def _write_raw_delivery_zip(entries: list[dict], zip_path: Path) -> None:
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    manifest = [item["manifest"] for item in entries]
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED, compresslevel=1) as zf:
        zf.writestr(
            RAW_MANIFEST_NAME,
            json.dumps(manifest, ensure_ascii=False, indent=2),
        )
        for item in entries:
            zf.write(item["path"], item["arcname"])


def create_raw_delivery_zip(db: Session, settings: Settings, *, force: bool = False) -> Path:
    file_count, max_mtime, entries = _raw_delivery_fingerprint(db, settings)
    cache_zip = settings.data_dir / RAW_DELIVERY_ZIP_NAME
    cache_meta = settings.data_dir / RAW_DELIVERY_ZIP_META_NAME

    cached = _read_delivery_cache_meta(cache_meta)
    if (
        not force
        and cache_zip.exists()
        and cached
        and cached.get("file_count") == file_count
        and cached.get("max_mtime") == max_mtime
    ):
        return cache_zip

    _write_raw_delivery_zip(entries, cache_zip)
    cache_meta.write_text(
        json.dumps(
            {
                "file_count": file_count,
                "max_mtime": max_mtime,
                "built_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return cache_zip


def _resolve_v2_pass_dir(settings: Settings, sample: Sample) -> Path | None:
    """定位新版交付 ZIP 用的 pass session 目录。"""
    storage_name = sample_storage_dir_name(sample.task_id, sample.session_id)
    paths = SourceSamplePaths.from_root(settings.samples_dir, sample.source_type)
    candidates = [paths.pass_dir / storage_name, paths.pass_dir / sample.session_id]

    backup_root = Path(sample.backup_dir) if sample.backup_dir else None
    if backup_root and backup_root.is_dir():
        backup_paths = SourceSamplePaths.from_root(backup_root, sample.source_type)
        candidates.extend(
            [
                backup_paths.pass_dir / storage_name,
                backup_paths.pass_dir / sample.session_id,
            ]
        )

    for path in candidates:
        if path.is_dir() and any(path.glob("*.json")):
            return path
    return None


def _build_qc_record_workbook(rows: list[dict], *, group: str) -> bytes:
    """生成质检提交记录.xlsx（分组 / 提交者 / sessionId / 对话场景）。"""
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl 依赖，请在服务器执行: uv sync") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ["分组", "提交者", "sessionId", "对话场景"]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)
    for idx, row in enumerate(rows, start=2):
        ws.cell(idx, 1, group)
        ws.cell(idx, 2, row["submitter"])
        ws.cell(idx, 3, row["session_storage_name"])
        ws.cell(idx, 4, row["scene_label"])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _v2_delivery_fingerprint(
    db: Session,
    settings: Settings,
    *,
    group: str,
) -> tuple[int, float, list[dict], str]:
    """统计新版交付条目与最新 mtime。"""
    rows = (
        db.query(Sample, Task, User.username)
        .join(Task, Sample.task_id == Task.id)
        .join(User, Sample.user_id == User.id)
        .order_by(Sample.id.asc())
        .all()
    )
    if not rows:
        raise FileNotFoundError("暂无已通过样本，请先完成至少一条通过样本")

    entries: list[dict] = []
    file_count = 0
    max_mtime = 0.0
    for sample, task, username in rows:
        pass_dir = _resolve_v2_pass_dir(settings, sample)
        if pass_dir is None:
            continue
        storage_name = sample_storage_dir_name(sample.task_id, sample.session_id)
        session_files: list[Path] = []
        for file_path in pass_dir.rglob("*"):
            if not file_path.is_file():
                continue
            session_files.append(file_path)
            file_count += 1
            max_mtime = max(max_mtime, file_path.stat().st_mtime)
        if not session_files:
            continue
        entries.append(
            {
                "source_type": sample.source_type,
                "session_storage_name": storage_name,
                "scene_label": task.scene_label or task.scene,
                "submitter": username,
                "pass_dir": pass_dir,
                "files": session_files,
            }
        )

    if not entries:
        raise FileNotFoundError("暂无可打包的 pass 样本目录，请检查 samples 目录")

    return file_count, max_mtime, entries, group.strip()


def _write_v2_delivery_zip(
    entries: list[dict],
    *,
    group: str,
    zip_path: Path,
) -> None:
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    xlsx_bytes = _build_qc_record_workbook(entries, group=group)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED, compresslevel=1) as zf:
        zf.writestr(QC_RECORD_XLSX_NAME, xlsx_bytes)
        for item in entries:
            for file_path in item["files"]:
                relative = file_path.relative_to(item["pass_dir"]).as_posix()
                arcname = f"{item['source_type']}/{item['session_storage_name']}/{relative}"
                zf.write(file_path, arcname)


def create_v2_delivery_zip(
    db: Session,
    settings: Settings,
    *,
    group: str,
    force: bool = False,
) -> Path:
    """
    新版交付 ZIP：
      质检提交记录.xlsx
      hermes/{taskId_sessionId}/...（仅 pass）
      openclaw/{taskId_sessionId}/...（仅 pass）
    提交者按样本实际用户名逐行写入。
    """
    if not group.strip():
        raise ValueError("分组不能为空")

    file_count, max_mtime, entries, group_value = _v2_delivery_fingerprint(
        db,
        settings,
        group=group,
    )
    cache_zip = settings.data_dir / V2_DELIVERY_ZIP_NAME
    cache_meta = settings.data_dir / V2_DELIVERY_ZIP_META_NAME

    cached = _read_delivery_cache_meta(cache_meta)
    if (
        not force
        and cache_zip.exists()
        and cached
        and cached.get("file_count") == file_count
        and cached.get("max_mtime") == max_mtime
        and cached.get("group") == group_value
        and cached.get("entry_count") == len(entries)
        and cached.get("submitter_mode") == "per_user"
    ):
        return cache_zip

    _write_v2_delivery_zip(
        entries,
        group=group_value,
        zip_path=cache_zip,
    )
    cache_meta.write_text(
        json.dumps(
            {
                "file_count": file_count,
                "max_mtime": max_mtime,
                "entry_count": len(entries),
                "group": group_value,
                "submitter_mode": "per_user",
                "built_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return cache_zip


def invalidate_delivery_zip_cache(settings: Settings) -> None:
    """有新样本入库后丢弃缓存，下次下载时重新打包。"""
    for name in (
        DELIVERY_ZIP_NAME,
        DELIVERY_ZIP_META_NAME,
        RAW_DELIVERY_ZIP_NAME,
        RAW_DELIVERY_ZIP_META_NAME,
        V2_DELIVERY_ZIP_NAME,
        V2_DELIVERY_ZIP_META_NAME,
    ):
        (settings.data_dir / name).unlink(missing_ok=True)
